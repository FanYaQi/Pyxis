import pandas as pd
import geopandas as gpd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
from openpyxl import load_workbook
import cartopy.crs as ccrs
import cartopy.feature as cfeature
from matplotlib.colors import LinearSegmentedColormap
from utils.path_util import DATA_PATH
from shapely import wkt

def load_opgee_results(filepath):
    wb = load_workbook(filepath, data_only=True)
    ws = wb['Results']
    num_fields = 269  # Updated to use a fixed number of fields
    field_names = [ws.cell(row=21, column=i).value for i in range(8, 8 + num_fields)]
    
    # Extract oil production and lifecycle GHG emissions data
    oil_production = [ws.cell(row=24, column=i).value for i in range(8, 8 + num_fields)]
    lifecycle_GHG_emissions = [ws.cell(row=193, column=i).value for i in range(8, 8 + num_fields)]
    
    # Create a DataFrame to store the extracted data
    data = {
        'Field name': field_names,
        'Oil Production': oil_production,
        'Lifecycle GHG Emissions': lifecycle_GHG_emissions
    }
    results_df = pd.DataFrame(data)
    return results_df

def load_basin_shapefile(filepath):
    basin_gdf = gpd.read_file(filepath)
    basin_gdf['basin_name'] = basin_gdf['name'].str.replace(r'_Mar$', '_Offshore', regex=True)
    basin_gdf['basin_name'] = basin_gdf['basin_name'].str.replace(r'_Terra$', '_Onshore', regex=True)
    merged_basins_gdf = basin_gdf.dissolve(by='basin_name').reset_index()
    return merged_basins_gdf

def load_pyxis_data(filepath):
    pyxis_data = pd.read_csv(filepath)
    pyxis_data['geometry'] = pyxis_data['geometry'].apply(wkt.loads)
    pyxis_fields_gdf = gpd.GeoDataFrame(pyxis_data, geometry='geometry', crs="EPSG:4326")
    return pyxis_fields_gdf

def load_flare_data(filepath):
    flare_data = pd.read_csv(filepath)
    return flare_data

def prepare_merged_with_basin_gdf(results_df, pyxis_fields_gdf, basin_gdf, flare_data):
    merged_gdf = pd.merge(results_df, pyxis_fields_gdf, on='Field name', how='left')
    merged_gdf = gpd.GeoDataFrame(merged_gdf, geometry='geometry', crs=pyxis_fields_gdf.crs)
    basin_gdf = basin_gdf.to_crs(merged_gdf.crs)
    merged_with_basin_gdf = gpd.sjoin(merged_gdf, basin_gdf, how='left', predicate='within')
    merged_with_basin_gdf = merged_with_basin_gdf[['Field name', 'Pyxis ID', 'Oil Production', 'Lifecycle GHG Emissions', 'basin_name', 'Gas-to-oil ratio (GOR)', 'Offshore']].merge(
        flare_data[['Pyxis ID', 'Flaring-to-oil ratio']], on='Pyxis ID', how='left')
    return merged_with_basin_gdf

def prepare_basin_avg(merged_with_basin_gdf):
    basin_avg = merged_with_basin_gdf.groupby('basin_name').apply(
        lambda x: pd.Series({
            'volume_weighted_CI': np.average(x['Lifecycle GHG Emissions'], weights=x['Oil Production']),
            'average_FOR': x['Flaring-to-oil ratio'].mean(),
            'total_production': x['Oil Production'].sum(),
            'offshore_status': x['Offshore'].mode()[0]  # Offshore/Onshore status based on the majority value
        })
    ).reset_index()
    basin_avg['average_FOR'].fillna(0, inplace=True)
    return basin_avg

def create_impact_plot(merged_with_basin_gdf, output_path):
    plt.rcParams['font.family'] = 'Helvetica'
    plt.rcParams['font.size'] = 12
    
    merged_with_basin_gdf['Impact'] = merged_with_basin_gdf['Oil Production'] * merged_with_basin_gdf['Lifecycle GHG Emissions']
    conditions = [
        (merged_with_basin_gdf['Gas-to-oil ratio (GOR)'] <= 100),
        (merged_with_basin_gdf['Gas-to-oil ratio (GOR)'] > 100) & (merged_with_basin_gdf['Gas-to-oil ratio (GOR)'] <= 500),
        (merged_with_basin_gdf['Gas-to-oil ratio (GOR)'] > 500) & (merged_with_basin_gdf['Gas-to-oil ratio (GOR)'] <= 1000),
        (merged_with_basin_gdf['Gas-to-oil ratio (GOR)'] > 1000) & (merged_with_basin_gdf['Gas-to-oil ratio (GOR)'] <= 5000),
        (merged_with_basin_gdf['Gas-to-oil ratio (GOR)'] > 5000) & (merged_with_basin_gdf['Gas-to-oil ratio (GOR)'] <= 10000),
        (merged_with_basin_gdf['Gas-to-oil ratio (GOR)'] > 10000)
    ]
    
    colors = ['red', 'orange', 'yellow', 'lightgreen', 'seagreen', 'darkgreen']
    merged_with_basin_gdf['color'] = np.select(conditions, colors)
    merged_with_basin_gdf = merged_with_basin_gdf.sort_values(by='Lifecycle GHG Emissions')
    
    merged_with_basin_gdf['width'] = merged_with_basin_gdf['Oil Production'] / merged_with_basin_gdf['Oil Production'].sum()
    merged_with_basin_gdf['cumulative_x'] = merged_with_basin_gdf['width'].cumsum() - merged_with_basin_gdf['width']
    
    volume_weighted_avg_CI = (merged_with_basin_gdf['Lifecycle GHG Emissions'] * merged_with_basin_gdf['Oil Production']).sum() / merged_with_basin_gdf['Oil Production'].sum()
    
    plt.figure(figsize=(12, 6), dpi=300)
    hatch_patterns = {
        'Santos': '///',
        'Campos': '...',
        'Others': ''
    }
    
    for i, row in merged_with_basin_gdf.iterrows():
        hatch = hatch_patterns.get(row['basin_name'], hatch_patterns['Others'])
        plt.bar(row['cumulative_x'], row['Lifecycle GHG Emissions'], color=row['color'], width=row['width'], edgecolor='k', alpha=0.6, align='edge', hatch=hatch)
    
    plt.axhline(y=volume_weighted_avg_CI, color='k', linestyle='--', linewidth=1, label=f'Volume-Weighted Avg CI: {volume_weighted_avg_CI:.2f}')
    plt.text(0.05, volume_weighted_avg_CI + 0.5, f'The volume-weighted average upstream CI: {volume_weighted_avg_CI:.2f} gCO$_2$eq/MJ',
             verticalalignment='bottom', horizontalalignment='left', color='black', fontsize=10)
    
    top_fields = merged_with_basin_gdf.nlargest(3, 'Impact')
    for i, row in top_fields.iterrows():
        plt.annotate(row['Field name'], xy=(row['cumulative_x'] + row['width'] / 2, row['Lifecycle GHG Emissions']),
                     xytext=(row['cumulative_x'] + row['width'] / 2, row['Lifecycle GHG Emissions'] + 5),
                     arrowprops=dict(facecolor='black', arrowstyle="->"), ha='center', fontsize=10, color='black')
    
    plt.ylim(0, 30)
    plt.xlim(0, 1)
    plt.xlabel('Cumulative Oil Production (%)')
    plt.ylabel('Lifecycle GHG Emissions (gCO$_2$eq/MJ)')
    plt.title('Lifecycle GHG Emissions per Field with GOR and Basin Categories')
    
    x_ticks = np.linspace(0, 1, num=6)
    x_labels = [f'{int(tick * 100)}%' for tick in x_ticks]
    plt.xticks(x_ticks, x_labels)
    
    legend_labels = [
        r'(0, 100]', r'(100, 500]', r'(500, 1000]', r'(1000, 5000]', r'(5000, 10000]', r'(10000, Inf)'
    ]
    legend_handles = [Patch(facecolor=col, label=label) for col, label in zip(colors, legend_labels)]
    
    hatch_legend = [
        Patch(facecolor='white', edgecolor='k', hatch='///', label='Santos'),
        Patch(facecolor='white', edgecolor='k', hatch='...', label='Campos'),
        Patch(facecolor='white', edgecolor='k', hatch='', label='Other Basins')
    ]
    
    plt.legend(handles=legend_handles + hatch_legend, title="GOR and Basin Categories", frameon=False)
    
    ax1 = plt.gca()
    ax2 = ax1.twiny()
    
    annual_production_million_bbl_d = merged_with_basin_gdf['Oil Production'].sum() / 1e3
    top_x_ticks = np.linspace(0, 1, num=6)
    top_x_labels = [f'{int(tick * annual_production_million_bbl_d)}' for tick in top_x_ticks]
    ax2.set_xlim(ax1.get_xlim())
    ax2.set_xticks(top_x_ticks)
    ax2.set_xticklabels(top_x_labels)
    ax2.set_xlabel("Cumulative Oil Production (thousand bbl/d)")
    ax2.tick_params(direction='in')
    
    plt.savefig(output_path, format='svg', dpi=300)
    plt.show()

def create_basin_plots(basin_gdf_mt, basin_avg, output_path):
    plt.rcParams['font.family'] = 'Helvetica'
    plt.rcParams['font.size'] = 13

    volume_weighted_avg_CI = basin_avg['volume_weighted_CI'].mean()

    fig, (ax1, ax2, ax3) = plt.subplots(nrows=1, ncols=3, figsize=(20, 10), dpi=300, gridspec_kw={'width_ratios': [1, 1, 2.5]})

    bars1 = ax1.barh(basin_avg['basin_name'], basin_avg['volume_weighted_CI'], color=['#B1040E' if status else "#006CB8" for status in basin_avg['offshore_status']], edgecolor='k')
    ax1.axvline(x=volume_weighted_avg_CI, color='black', linestyle='--', linewidth=1, label=f'Avg CI: {volume_weighted_avg_CI:.2f}')
    ax1.set_xlabel('Volume-Weighted Average CI (gCO$_2$eq/MJ)')
    ax1.invert_yaxis()

    ax2.barh(basin_avg['basin_name'], basin_avg['total_production'], color='#008566', edgecolor='k')
    ax2.set_xlabel('Oil Production (million bbl/d)')
    ax2.invert_yaxis()

    if 'ax3' in locals():
        ax3.remove()
    ax3 = plt.subplot(133, projection=ccrs.PlateCarree())
    ax3.set_extent([-75, -30, -35, 10])
    ax3.add_feature(cfeature.LAND, zorder=0, edgecolor='black', facecolor='white', alpha=0.3)
    ax3.add_feature(cfeature.COASTLINE)
    ax3.add_feature(cfeature.BORDERS, linestyle=':')

    cmap = LinearSegmentedColormap.from_list("custom_cmap", ["#e9c716", '#bc272d'])
    norm = plt.Normalize(basin_avg['average_FOR'].min(), basin_avg['average_FOR'].max())

    basin_gdf_mt.plot(column='average_FOR', cmap=cmap, linewidth=0.8, ax=ax3, edgecolor='1', legend=False, zorder=2)
    for idx, row in basin_gdf_mt.iterrows():
        ax3.text(row.geometry.centroid.x, row.geometry.centroid.y, row['basin_name'], fontsize=8, ha='center', zorder=3)

    sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm)
    sm.set_array([])
    cbar = fig.colorbar(sm, ax=ax3, orientation='vertical', fraction=0.05, pad=0.05)
    cbar.set_label('Average FOR (scf/bbl)')

    plt.subplots_adjust(wspace=0.1)
    plt.savefig(output_path, format='svg', dpi=300)
    plt.show()

def main():
    excel_path = f'{DATA_PATH}/OPGEE_model/OPGEE_3.0b_BETA_BR.xlsm'
    basin_shapefile = f'{DATA_PATH}/br_geodata/br_basin_shp/bacias_gishub_db.shp'
    pyxis_data_path = f'{DATA_PATH}/br_geodata/merged_pyxis_field_info_table_filtered.csv'
    flare_data_path = f'{DATA_PATH}/br_geodata/merged_pyxis_field_info_with_flare.csv'
    impact_plot_output_path = f'{DATA_PATH}/br_geodata/plots/Field_CI_Plot_GOR_Basin_old.svg'
    basin_plot_output_path = f'{DATA_PATH}/br_geodata/plots/Basin_CI_Plot_with_production_and_for_old.svg'

    results_df = load_opgee_results(excel_path)
    basin_gdf = load_basin_shapefile(basin_shapefile)
    pyxis_fields_gdf = load_pyxis_data(pyxis_data_path)
    flare_data = load_flare_data(flare_data_path)

    # Prepare merged_with_basin_gdf and basin_avg using separate functions
    merged_with_basin_gdf = prepare_merged_with_basin_gdf(results_df, pyxis_fields_gdf, basin_gdf, flare_data)
    basin_avg = prepare_basin_avg(merged_with_basin_gdf)

    # Create plots
    create_impact_plot(merged_with_basin_gdf, impact_plot_output_path)
    create_basin_plots(basin_gdf, basin_avg, basin_plot_output_path)

if __name__ == "__main__":
    main()