from openpyxl import load_workbook
import pandas as pd


def load_opgee_results(filepath, num_fields):
    wb = load_workbook(filepath, data_only=True)
    ws = wb["Results"]
    field_names = [ws.cell(row=21, column=i).value for i in range(8, num_fields + 8)]

    # Extract oil production and lifecycle GHG emissions data
    oil_production = [ws.cell(row=24, column=i).value for i in range(8, num_fields + 8)]
    lifecycle_GHG_emissions = [
        ws.cell(row=193, column=i).value for i in range(8, num_fields + 8)
    ]

    # Create a DataFrame to store the extracted data
    data = {
        "Field name": field_names,
        "Oil Production": oil_production,
        "Lifecycle GHG Emissions": lifecycle_GHG_emissions,
    }
    results_df = pd.DataFrame(data)
    return results_df
