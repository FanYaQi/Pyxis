import pandas as pd
from openpyxl import load_workbook
from utils.path_util import DATA_PATH


def load_and_filter_fields(filepath):
    fields = pd.read_csv(filepath)
    oil_fields = fields[fields["Function unit"] == "oil"]
    return oil_fields[
        oil_fields["Oil production volume"].notna()
        & (oil_fields["Oil production volume"] > 0)
    ]


def clean_data(input_data):
    # Map the "Offshore" column to "Offshore?"
    input_data = input_data.rename(columns={"Offshore": "Offshore?"})

    # Set "Number of producing wells" and "Number of water injecting wells" to null if the number is smaller than 1
    input_data.loc[
        input_data["Number of producing wells"] < 1, "Number of producing wells"
    ] = None
    input_data.loc[
        input_data["Number of water injecting wells"] < 1,
        "Number of water injecting wells",
    ] = None

    # Set "API gravity (oil at standard pressure and temperature, or 'dead oil')" to null if outside 3-88 range
    api_column = 'API gravity (oil at standard pressure and temperature, or "dead oil")'
    input_data.loc[
        (input_data[api_column] < 3) | (input_data[api_column] > 88), api_column
    ] = None

    # # Set "Gas-to-oil ratio (GOR)" to null if it is larger than 34000 (OPGEE report value)
    # input_data.loc[input_data['Gas-to-oil ratio (GOR)'] > 34000, 'Gas-to-oil ratio (GOR)'] = None

    # Adjust "Flaring-to-oil ratio" if it's larger than "Gas-to-oil ratio (GOR)"
    input_data.loc[
        input_data["Flaring-to-oil ratio"] > input_data["Gas-to-oil ratio (GOR)"],
        "Flaring-to-oil ratio",
    ] = input_data["Gas-to-oil ratio (GOR)"]

    # Add "Brazil" to the "Field location (Country)" column for all fields
    input_data["Field location (Country)"] = "Brazil"

    return input_data


def populate_excel_template(input_data, template_path, output_path):
    # Load the Excel template and select the "Inputs" sheet
    wb = load_workbook(template_path, keep_vba=True)
    ws = wb["Inputs"]

    start_column = 8
    column_offset_counter = 0

    # Loop over each field in the input data
    for i, row in input_data.iterrows():
        column_offset = start_column + column_offset_counter

        # Populate the Excel template with the data
        for column_name in input_data.columns:
            found = False
            # Search for the row in columns A, B, and C
            for col in ["A", "B", "C"]:
                for cell in ws[col]:
                    if cell.value == column_name:
                        ws.cell(
                            row=cell.row, column=column_offset, value=row[column_name]
                        )
                        found = True
                        break
                if found:
                    break

        column_offset_counter += 1

    # Save the populated Excel file
    wb.save(output_path)


def main():
    # File paths
    data_filepath = (
        f"{DATA_PATH}/br_geodata/merged_pyxis_field_info_with_flare_withwm.csv"
    )
    excel_template_path = f"{DATA_PATH}/OPGEE_model/OPGEE_3.0c_BETA.xlsm"
    output_path = f"{DATA_PATH}/OPGEE_model/OPGEE_3.0c_BETA_BR_withwm.xlsm"

    # Load and filter fields
    oil_fields = load_and_filter_fields(data_filepath)

    # Clean data
    cleaned_data = clean_data(oil_fields)

    # Populate the Excel template
    populate_excel_template(cleaned_data, excel_template_path, output_path)


def main_wowm():
    # File paths
    data_filepath = (
        f"{DATA_PATH}/br_geodata/merged_pyxis_field_info_with_flare_wowm.csv"
    )
    excel_template_path = f"{DATA_PATH}/OPGEE_model/OPGEE_3.0c_BETA.xlsm"
    output_path = f"{DATA_PATH}/OPGEE_model/OPGEE_3.0c_BETA_BR_wowm.xlsm"

    # Load and filter fields
    oil_fields = load_and_filter_fields(data_filepath)

    # Clean data
    cleaned_data = clean_data(oil_fields)

    # Populate the Excel template
    populate_excel_template(cleaned_data, excel_template_path, output_path)


# Call main function
if __name__ == "__main__":
    main_wowm()
